"""
excel_writer.py
---------------
Excel (.xlsx) export.

One writer serves every report type: it takes the **same headers and rows that
were just written into the Word document** - never the PDF, never a second
parse - so the two files cannot drift apart.  Any future report (Motor
Starting, Arc Flash, cable schedules) gets its Excel export for free, because
:class:`reports.ReportResult` already carries the table.

The sheet is styled to read like an engineering schedule rather than a data
dump: bold header band, borders, frozen header, sensible column widths,
wrapped IDs, and numeric cells written as *numbers* with the same number of
decimals the Word document shows.
"""

from __future__ import annotations

import io
import os
import re
from typing import Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

#: Matches a plain decimal number, so it can be stored as a number not a string.
_NUMBER_RE = re.compile(r"^-?\d+(\.\d+)?$")

FONT_NAME = "Arial"
HEADER_FILL = "FF524689"   # the purple of the Word templates
HEADER_TEXT = "FFFFFFFF"
GRID_COLOUR = "FF7F7F7F"

#: Columns wider than this wrap instead of stretching the sheet.
MAX_WIDTH = 34
MIN_WIDTH = 9


def _thin_border() -> Border:
    side = Side(style="thin", color=GRID_COLOUR)
    return Border(left=side, right=side, top=side, bottom=side)


def _as_cell_value(text: str) -> tuple[object, Optional[str]]:
    """
    Convert one formatted string into ``(value, number_format)``.

    ``"19.09"`` becomes the number 19.09 shown as ``0.00`` - identical on
    screen, but sortable and usable in further calculations.  Anything else
    (``"132 kV"``, ``"ACCEPTABLE"``, ``""``) stays text.
    """
    text = "" if text is None else str(text)
    stripped = text.strip()
    if not _NUMBER_RE.match(stripped):
        return text, None

    decimals = len(stripped.split(".")[1]) if "." in stripped else 0
    number_format = "0." + "0" * decimals if decimals else "0"
    return float(stripped) if decimals else int(stripped), number_format


def build_workbook(
    headers: Sequence[str],
    rows: Sequence[Sequence[str]],
    sheet_name: str = "Report",
    title: Optional[str] = None,
) -> Workbook:
    """
    Build a workbook holding exactly *headers* and *rows*.

    Parameters mirror the Word table: same column order, same row order, same
    values.  No calculation and no reordering happens here.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (sheet_name or "Report")[:31]

    border = _thin_border()
    header_font = Font(name=FONT_NAME, bold=True, color=HEADER_TEXT, size=10)
    body_font = Font(name=FONT_NAME, size=10)
    fill = PatternFill("solid", fgColor=HEADER_FILL)

    first_data_row = 1
    if title:
        cell = sheet.cell(row=1, column=1, value=title)
        cell.font = Font(name=FONT_NAME, bold=True, size=12)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        first_data_row = 3

    # -- header row --------------------------------------------------------- #
    header_row = first_data_row
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=index, value=str(header))
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 30

    # -- data --------------------------------------------------------------- #
    for row_offset, row in enumerate(rows, start=header_row + 1):
        for index, raw in enumerate(row, start=1):
            value, number_format = _as_cell_value(raw)
            cell = sheet.cell(row=row_offset, column=index, value=value)
            cell.font = body_font
            cell.border = border
            if number_format:
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif index == 1:
                # Bus / Switchgear IDs: left aligned, wrapped when long.
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # -- widths, freeze pane, filter ---------------------------------------- #
    for index, header in enumerate(headers, start=1):
        longest = max(
            [len(str(header).split("\n")[0])]
            + [len(str(row[index - 1])) for row in rows if index - 1 < len(row)]
            or [MIN_WIDTH]
        )
        width = min(max(longest + 3, MIN_WIDTH), MAX_WIDTH)
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    if rows:
        sheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"
        )
    sheet.sheet_view.showGridLines = False
    sheet.print_title_rows = f"{header_row}:{header_row}"
    # Print as a schedule: landscape, one page wide, header repeated.
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    return workbook


def build_workbook_bytes(
    headers: Sequence[str],
    rows: Sequence[Sequence[str]],
    sheet_name: str = "Report",
    title: Optional[str] = None,
) -> bytes:
    """The workbook as raw ``.xlsx`` bytes (downloads / in-memory use)."""
    buffer = io.BytesIO()
    build_workbook(headers, rows, sheet_name, title).save(buffer)
    return buffer.getvalue()


def excel_path_for(docx_path: str) -> str:
    """``.../Bus Report.docx`` -> ``.../Bus Report.xlsx`` (same folder, same stem)."""
    base, _ = os.path.splitext(docx_path)
    return base + ".xlsx"


def write_excel(data: bytes, output_path: str) -> str:
    """Save prepared workbook bytes, creating the folder if needed."""
    folder = os.path.dirname(os.path.abspath(output_path))
    if folder:
        os.makedirs(folder, exist_ok=True)
    with open(output_path, "wb") as handle:
        handle.write(data)
    return output_path
